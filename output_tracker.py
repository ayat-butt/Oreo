import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
from datetime import date
from dateutil.relativedelta import relativedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

today = date(2026, 4, 2)

all_employees = [
    # Oct 2025
    ("Abdul Waheed",        "Regional Manager",               "Program",            "Niete", date(2025,10,1),  "Contract/Addendum request email on record",   "29 Sep 2025", "Not sent — 91 days overdue"),
    ("Hashir Hussain",      "Regional Manager",               "Program",            "Niete", date(2025,10,6),  "Experience letter request email Feb 2026",    "06 Oct 2025", "Not sent — 86 days overdue"),
    ("Javariya Mufarrakh",  "Sr Manager - People Ops",        "P&C",                "OPL",   date(2025,10,10), "Recent meeting email confirms active Apr 2026","10 Oct 2025", "Not sent — 82 days overdue"),
    ("Toseef Ur Rehman",    "CPD Coach",                      "Program",            "Niete", date(2025,10,20), "Contract Drafting & Onboarding email on record","12 Nov 2025","Not sent — 72 days overdue"),
    ("Ali Sipra",           "COO",                            "Leadership Team",    "OPL",   date(2025,10,20), "Welcome to Taleemabad - Contracts & Next Steps","19 Nov 2025","Not sent — 72 days overdue"),
    ("Zainab Zaheer",       "CPD Coach",                      "Program",            "Niete", date(2025,10,21), "Congratulations - NIETE Coach email on record","21 Dec 2025", "Not sent — 71 days overdue"),
    # Nov 2025
    ("Samra Tariq",          "Senior Finance Executive",       "Finance",            "OWT",   date(2025,11,10), "Welcome email on record",                    "10 Nov 2025", "Not sent — 51 days overdue"),
    ("Saleh Muhammad",       "OPL Staff",                      "OPL",                "OPL",   date(2025,11,1),  "Insurance email confirms active",             "01 Nov 2025", "Not sent — 60 days overdue"),
    ("Asma Zaheer",          "Regional Manager",               "Program",            "Niete", date(2025,11,17), "Office invite activity confirms active",       "17 Nov 2025", "Not sent — 44 days overdue"),
    ("Javeria Khalil",       "Coach",                          "Program",            "Niete", date(2025,11,27), "Office invite activity confirms active",       "27 Nov 2025", "Not sent — 34 days overdue"),
    # Dec 2025
    ("Aleeha Noor",          "Coach",                          "Program",            "Niete", date(2025,12,1),  "Office invite activity confirms active",       "01 Dec 2025", "Not sent — 32 days overdue"),
    ("Abdullah Durrani",     "Program Coordinator",            "Niete Balochistan",  "Niete", date(2025,12,4),  "Congratulations email on record",             "04 Dec 2025", "Not sent — 29 days overdue"),
    ("Taloot Ahmad Malik",   "Deputy Manager - Admin Ops",     "Admin & Operations", "OPL",   date(2025,12,23), "Active - recent emails confirm working",       "23 Dec 2025", "Not sent — 10 days overdue"),
    # Jan 2026
    ("M. Saim",              "Full Stack Developer",           "Engineering",        "OPL",   date(2026,1,1),   "Welcome to Taleemabad - Full Stack Developer", "17 Dec 2025", "Not sent — OVERDUE"),
    ("Muhammad Muzzamil Patel","Sr. Manager, Impact & Policy", "Data & Impact",      "OPL",   date(2026,1,8),   "Welcome to Taleemabad - Sr. Manager",          "12 Dec 2025", "Not sent — due 08 Apr 2026"),
    ("Saaim Asif",           "Coach",                          "Niete ICT",          "Niete", date(2026,1,13),  "Offer Letter CPD Coach | Saaim Asif",          "12 Jan 2026", "Not sent — due 13 Apr 2026"),
    ("Zarrish Ahmed",        "Growth Fellow",                  "Growth",             "OPL",   date(2026,1,26),  "Congratulations - Growth Fellow",              "21 Jan 2026", "Not sent — due 26 Apr 2026"),
    # Feb 2026
    ("Meerab Din",           "Coach",                          "Niete ICT",          "Niete", date(2026,2,2),   "Congratulations - NIETE Coach",                "21 Jan 2026", "Not sent — due 02 May 2026"),
    ("Usman Mughal",         "Sr. Manager Engineering",        "Engineering",        "OPL",   date(2026,2,16),  "Welcome to Taleemabad - Sr. Manager Engg",     "11 Feb 2026", "Not sent — due 16 May 2026"),
    ("Fakhr Ul Islam",       "Coach",                          "Niete ICT",          "Niete", date(2026,2,25),  "Congratulations - NIETE Coach",                "23 Feb 2026", "Not sent — due 25 May 2026"),
    ("Shoaib ud Din",        "Full Stack Developer",           "Engineering",        "OPL",   date(2026,2,26),  "Welcome to Taleemabad - Full Stack Developer", "18 Feb 2026", "Not sent — due 26 May 2026"),
    # Mar 2026
    ("Bibi Raheela",         "Training Manager",               "Niete RWP",          "Niete", date(2026,3,9),   "Congratulations - Training Manager NIETE",     "05 Mar 2026", "Not sent — due 09 Jun 2026"),
    ("Mehwish Bibi",         "Baby Sitter / Nanny",            "OWT",                "OWT",   date(2026,3,24),  "Contract sent via Oreo to Taloot (I-10)",       "01 Apr 2026", "Not sent — due 24 Jun 2026"),
]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Probation Tracker"

title_font   = Font(name="Calibri", bold=True,   color="FFFFFF", size=13)
hdr_font     = Font(name="Calibri", bold=True,   color="FFFFFF", size=10)
normal_font  = Font(name="Calibri",              size=10)
bold_font    = Font(name="Calibri", bold=True,   size=10)
white_bold   = Font(name="Calibri", bold=True,   color="FFFFFF", size=10)
red_bold     = Font(name="Calibri", bold=True,   color="C00000", size=10)
orange_bold  = Font(name="Calibri", bold=True,   color="FF8C00", size=10)
italic_font  = Font(name="Calibri", italic=True, color="7F7F7F", size=9)

hdr_fill     = PatternFill("solid", fgColor="1F4E79")
overdue_fill = PatternFill("solid", fgColor="7B0000")
done_fill    = PatternFill("solid", fgColor="C00000")
third_fill   = PatternFill("solid", fgColor="FF8C00")
second_fill  = PatternFill("solid", fgColor="2E75B6")
first_fill   = PatternFill("solid", fgColor="70AD47")
alert_fill   = PatternFill("solid", fgColor="FFF2CC")
section_fill = PatternFill("solid", fgColor="D6DCE4")

center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
thin   = Side(style="thin", color="BFBFBF")
bdr    = Border(left=thin, right=thin, top=thin, bottom=thin)

# Title
ws.merge_cells("A1:L1")
ws["A1"] = "Probation Status Report — April 2026  |  Active Employees Only  |  Verified: Sheets + Gmail"
ws["A1"].font = title_font
ws["A1"].fill = hdr_fill
ws["A1"].alignment = center
ws.row_dimensions[1].height = 32

# Note
ws.merge_cells("A2:L2")
ws["A2"] = ("Cross-checked: Oct 2025, Nov 2025, Dec 2025, Jan-Mar 2026 sheets + Gmail. "
            "Removed: Minha Khan (left 16 Dec 2025), Jarrar Ali Khan (left 1 Mar 2026), "
            "Ammad Rasheed (left 7 Jan 2026), Hassan Shahzad (left 6 Nov 2025), "
            "Zainab Fatima (left 22 Feb 2026), and 6 Niete Balochistan staff (left 28 Feb 2026).")
ws["A2"].font = italic_font
ws["A2"].alignment = left
ws.row_dimensions[2].height = 20

# Headers
headers = ["#", "Name", "Designation", "Department", "Entity",
           "Date of Joining", "Probation End Date", "Days Overdue / Left",
           "Status", "Contract on Record", "Contract Date", "Probation Closure Status"]
ws.append(headers)
ws.row_dimensions[3].height = 22
for col in range(1, 13):
    c = ws.cell(row=3, column=col)
    c.font = hdr_font; c.fill = hdr_fill; c.alignment = center; c.border = bdr

sections = [
    ("October 2025 Joiners",  all_employees[0:6]),
    ("November 2025 Joiners", all_employees[6:10]),
    ("December 2025 Joiners", all_employees[10:13]),
    ("January 2026 Joiners",  all_employees[13:17]),
    ("February 2026 Joiners", all_employees[17:21]),
    ("March 2026 Joiners",    all_employees[21:23]),
]

row_num = 4
sr = 1
for section_title, employees in sections:
    ws.merge_cells(f"A{row_num}:L{row_num}")
    c = ws.cell(row=row_num, column=1, value=section_title)
    c.font = Font(name="Calibri", bold=True, color="1F4E79", size=10)
    c.fill = section_fill; c.alignment = left; c.border = bdr
    ws.row_dimensions[row_num].height = 16
    row_num += 1

    for name, desig, dept, entity, joined, contract_ref, contract_date, closure in employees:
        prob_end  = joined + relativedelta(months=3)
        days_left = (prob_end - today).days

        if days_left <= 0:
            overdue_days = abs(days_left)
            status = f"Completed — {overdue_days}d overdue"
            s_fill = overdue_fill if overdue_days > 14 else done_fill
            days_str = f"-{overdue_days} days"
        elif days_left <= 30:
            status = "In 3rd Month — Ending Soon"; s_fill = third_fill; days_str = f"{days_left} days"
        elif days_left <= 60:
            status = "In 2nd Month — Upcoming";    s_fill = second_fill; days_str = f"{days_left} days"
        else:
            status = "In 1st Month";               s_fill = first_fill;  days_str = f"{days_left} days"

        ws.append([sr, name, desig, dept, entity,
                   joined.strftime("%d %b %Y"), prob_end.strftime("%d %b %Y"),
                   days_str, status, contract_ref, contract_date, closure])

        for col in range(1, 13):
            c = ws.cell(row=row_num, column=col)
            c.border = bdr
            c.alignment = center if col in (1, 5, 6, 7, 8, 11) else left
            if col == 2:    c.font = bold_font
            elif col == 9:  c.font = white_bold; c.fill = s_fill
            elif col == 12:
                c.fill = alert_fill
                c.font = red_bold if "overdue" in closure.lower() else (orange_bold if days_left <= 30 else normal_font)
            else:           c.font = normal_font
        ws.row_dimensions[row_num].height = 18
        row_num += 1
        sr += 1

# Column widths
for i, w in enumerate([3, 26, 30, 20, 7, 15, 16, 14, 28, 40, 13, 30], 1):
    ws.column_dimensions[get_column_letter(i)].width = w

wb.save("output/Probation_Tracker_April_2026.xlsx")
print(f"Done. {sr-1} employees in tracker.")
